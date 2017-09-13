from openpyxl import Workbook
import re, string

def ColumnSort(massiveString):
    splits = re.split('(Facility)',massiveString)
    del splits[0]
    newSplits = []
    num = 0
    while num < len(splits):
        newSplits.append(splits[num]+splits[num+1])
        num += 2
    splits = [record for record in newSplits]
    newSplits = []
    for record in splits:
        record.rstrip('\n')
        newSplits.append(record.split('\n'))
    splits = []
    for record in newSplits:
        appendList = []
        for variable in record:
            if variable != ['']:
                appendList.append(variable)
        splits.append(appendList)
    tupleArr = []
    for record in splits:
        recordTraitTupleArr = []
        for trait in record:
            traitSplit = trait.split('\t')
            if len(traitSplit) == 2:
                recordTraitTupleArr.append(traitSplit)
        tupleArr.append(recordTraitTupleArr)
    return tupleArr

def importData(columnArrOfDict):
    wb = Workbook()
    ws = wb.active
    usedTraits = []
    listOfLetters = list(string.ascii_uppercase)
    columnNum = 0
    letterTitleDict = dict()
    for record in columnArrOfDict:
        for trait in record:
            if trait[0] not in usedTraits:
                cellStr = str(listOfLetters[columnNum])+'1'
                ws[cellStr] = trait[0]
                letterTitleDict[trait[0]] = listOfLetters[columnNum]
                columnNum += 1
                usedTraits.append(trait[0])
    rowNum = 2
    for record in columnArrOfDict:
        for trait in record:
            cellStr = str(letterTitleDict[trait[0]]) + str(rowNum)
            ws[cellStr] = trait[1]
        rowNum += 1
    wb.save('C:\Openpyxl\massiveStr.xlsx')
    

if __name__ == '__main__':
    massiveString = """Facility :	Active Recycling Corp.
DEP ID :	133660
Phone:	908-725-8484
City :	Somerville
Zip :	08876
Location :	15 Polhemus Lane
County :	Somerset
Municipality :	Bridgewater Twp.
Facility :	Atlantic Cty UA
DEP ID :	141531
Phone:	609-272-6950
City :	Pleasantville
Zip :	08232
Location :	6700 Delilah Road
County :	Atlantic
Municipality :	Egg Harbor
Facility :	B & J Recycling
DEP ID :	131834
Phone:	609-652-2434
City :	Galloway
Zip :	08205
Location :	141 S. Old Port Road
County :	Atlantic
Municipality :	Galloway
Facility :	Bayshore Recycling Corp
DEP ID :	132397
Phone:	732-738-6000
City :	Keasby
Zip :	08832
Location :	75 Crows Mill Road
County :	Middlesex
Municipality :	Keasby
Facility :	Burlington County Freeholders
DEP ID :	131962
Phone:	609-499-1001
City :	Mt. Holly
Zip :	08060
Location :	Burlington-Columbus Road
County :	Burlington
Municipality :	Florence
Facility :	Cape May County MUA
DEP ID :	133646
Phone:	609-465-9026
City :	CMCH
Zip :	08210
Location :	2050 Dennisville-Petersburg Road
County :	Cape May
Municipality :	Woodbine
Facility :	County Conservation
DEP ID :	132201
Phone:	856-227-6900
City :	Washington Twp
Zip :	08080
Location :	212 Blackwood Barnsboro Road
County :	Gloucester
Municipality :	Washington
Facility :	Durable Recycling
DEP ID :	160846
Phone:	201-437-0703
City :	Bayonne
Zip :	07002
Location :	195 East 22nd Street
County :	Hudson
Municipality :	Bayonne
Facility :	Evergreen Recycling Solutions, LLC
DEP ID :	499991
Phone:	973-242-3030
City :	Westfield
Zip :	07090
Location :	110 Evergreen Avenue
County :	Essex
Municipality :	Newark
Facility :	Freehold Cartage, Inc.
DEP ID :	132412
Phone:	732-462-1001
City :	Freehold
Zip :	07728
Location :	825 Highway 33 East
County :	Monmouth
Municipality :	Freehold
Facility :	Garden State Landscape Products, LLC
DEP ID :	270401
Phone:	908-638-0199
City :	Oxford
Zip :	07863
Location :	2026 Rt. 31 & Buffalo Hollow Road
County :	Hunterdon
Municipality :	Lebanon
Facility :	Generated Materials LLC
DEP ID :	132324
Phone:	732-985-3370
City :	Edison
Zip :	08817
Location :	327 Meadow Road
County :	Middlesex
Municipality :	Edison
Facility :	Gloucester County Mulch Factory LLC
DEP ID :	132203
Phone:	856-589-1501
City :	Turnersville
Zip :	08012
Location :	120 Sewell Road
County :	Gloucester
Municipality :	Washington
Facility :	GreenRock Recycling, LLC
DEP ID :	591306
Phone:	908-735-9726
City :	Union
Zip :	08827
Location :	5 & 6 Frontage Road
County :	Hunterdon
Municipality :	Union
Facility :	Grinnell Recycling, Inc.
DEP ID :	132685
Phone:	973-383-9300
City :	Sparta
Zip :	07871
Location :	482 Houses Corner Road
County :	Sussex
Municipality :	Sparta
Facility :	Hamilton Township
DEP ID :	132279
Phone:	609-890-3506
City :	Hamilton
Zip :	08650
Location :	1360 Kuser Road
County :	Mercer
Municipality :	Hamilton
Facility :	Iaconelli Contracting Co., Inc.
DEP ID :	131845
Phone:	609-645-2165
City :	Pleasantville
Zip :	08232
Location :	977 Mill Road
County :	Atlantic
Municipality :	Pleasantville
Facility :	J. Manzo Recycling
DEP ID :	132430
Phone:	732-946-7100
City :	Matawan
Zip :	07747
Location :	55 Highway 34
County :	Monmouth
Municipality :	Marlboro
Facility :	J.H. Reid
DEP ID :	132339
Phone:	732-752-4050
City :	S. Plainfield
Zip :	07080
Location :	172 Baekeland Avenue
County :	Middlesex
Municipality :	Middlesex
Facility :	Lawrence Township
DEP ID :	132290
Phone:	609-844-7137
City :	Lawrenceville
Zip :	08648
Location :	3701 Princeton Pike
County :	Mercer
Municipality :	Lawrence Township
Facility :	Lertch Recycling Company, Inc.
DEP ID :	132460
Phone:	732-681-0206
City :	Wall
Zip :	07719
Location :	5115 Belmar Boulevard
County :	Monmouth
Municipality :	Wall Twp.
Facility :	Mazza & Sons
DEP ID :	132440
Phone:	732-922-9292
City :	Oakhurst
Zip :	07755
Location :	3230 Shafto Road
County :	Monmouth
Municipality :	Tinton Falls
Facility :	Mercer Group International
DEP ID :	132273
Phone:	609-631-8230
City :	Trenton
Zip :	08625
Location :	4 Beakes Street
County :	Mercer
Municipality :	Ewing
Facility :	Mid-Jersey Mulch Prdts.
DEP ID :	132289
Phone:	609-588-8225
City :	Lawrenceville
Zip :	08648
Location :	227 Bakers Basin Road
County :	Mercer
Municipality :	Lawrenceville
Facility :	Miele Sanitation
DEP ID :	197012
Phone:	201-768-3818
City :	Closter
Zip :	07624
Location :	60 Railroad Avenue
County :	Bergen
Municipality :	Closter
Facility :	Mimlitsch
DEP ID :	131954
Phone:	856-985-0412
City :	Evesham
Zip :	08053
Location :	151 New Road
County :	Burlington
Municipality :	Evesham
Facility :	Nature's Choice - Upper Deerfield
DEP ID :	132084
Phone:	201-531-0260
City :	Union
Zip :	07083
Location :	1310 Route 77
County :	Cumberland
Municipality :	Bridgeton
Facility :	North Jersey Wood Products, LLC
DEP ID :	474892
Phone:	973-798-6071
City :	Fairfield
Zip :	07004
Location :	123 Duffield Avenue
County :	Hudson
Municipality :	Jersey City
Facility :	Old Cape Recycling
DEP ID :	131824
Phone:	609-926-6420
City :	Egg Harbor Twp.
Zip :	08234
Location :	3025 Ocean Heights Avenue
County :	Atlantic
Municipality :	Egg Harbor
Facility :	Puggi, Anthony
DEP ID :	133470
Phone:	609-926-6991
City :	Mays Landing
Zip :	08330
Location :	6151 Mill Road
County :	Atlantic
Municipality :	Egg Harbor
Facility :	Raritan Valley Recycling
DEP ID :	132262
Phone:	908-782-4004
City :	Whitehouse Station
Zip :	08889
Location :	79 River Road
County :	Hunterdon
Municipality :	Raritan Twp.
Facility :	Reclamation Technology, Inc.
DEP ID :	132331
Phone:	908-769-0600
City :	Bernardsville
Zip :	07924
Location :	3200 Bordentown Avenue
County :	Middlesex
Municipality :	Old Bridge
Facility :	Recycling of Central Jersey, LLC
DEP ID :	132543
Phone:	732-323-0226
City :	Jackson
Zip :	08527
Location :	577 South Hope Chapel Road
County :	Ocean
Municipality :	Jackson
Facility :	Reliable Paper
DEP ID :	134654
Phone:	201-333-5244
City :	Jersey City
Zip :	07304
Location :	1 Caven Point Avenue
County :	Hudson
Municipality :	Jersey City
Facility :	Reliable Wood Products, Inc.
DEP ID :	132312
Phone:	201-333-5244
City :	Jersey City
Zip :	07304
Location :	234 Broadway Road
County :	Middlesex
Municipality :	South Brunswick
Facility :	RER Supply, LLC
DEP ID :	567091
Phone:	973-616-6654
City :	Riverdale
Zip :	07457
Location :	4 South Corporate Drive
County :	Morris
Municipality :	Riverdale
Facility :	Resource Engineering, LLC
DEP ID :	543881
Phone:	
City :	Farmingdale
Zip :	07727
Location :	34 Randolph Road
County :	Monmouth
Municipality :	Howell
Facility :	River Front Recycling & Aggregate, LLC
DEP ID :	132011
Phone:	856-966-1100
City :	Camden
Zip :	08105
Location :	1301 North 26th Street
County :	Camden
Municipality :	Camden
Facility :	Silva Construction & Demolition Inc.
DEP ID :	502757
Phone:	973-483-3792
City :	Newark
Zip :	07104
Location :	100 Riverside Avenue
County :	Essex
Municipality :	Newark
Facility :	Skytop Recycling, Inc
DEP ID :	132603
Phone:	973-728-0001
City :	West Milford
Zip :	07480
Location :	80 Airport Road
County :	Passaic
Municipality :	West Milford
Facility :	South Jersey Agr. Prod.
DEP ID :	133658
Phone:	856-358-0990
City :	Elmer
Zip :	08318
Location :	395 Route 77 South
County :	Salem
Municipality :	Upper Pittsgrove
Facility :	T. Fiore Recycling Corp.
DEP ID :	132148
Phone:	973-589-3366
City :	Newark
Zip :	07105
Location :	411 Wilson Avenue
County :	Essex
Municipality :	Newark
Facility :	Thompson's Paving Center
DEP ID :	478179
Phone:	856-358-3452
City :	Elmer
Zip :	08318
Location :	1020 Buckshutem Road
County :	Cumberland
Municipality :	Fairfield Twp
Facility :	Tilcon New York, Inc.
DEP ID :	132238
Phone:	201-997-0220
City :	Kearny
Zip :	07032
Location :	411 Bergen Avenue
County :	Hudson
Municipality :	Kearny
Facility :	Vinch Recycling
DEP ID :	132272
Phone:	609-393-0200
City :	Trenton
Zip :	08638
Location :	1607 North Olden Avenue
County :	Mercer
Municipality :	Trenton
Facility :	Vollers Excavating, Inc.
DEP ID :	132638
Phone:	908-725-1026
City :	North Branch
Zip :	08876
Location :	3311 U.S. Route 22 South
County :	Somerset
Municipality :	Branchburg
Facility :	W.P.A.R., Inc.
DEP ID :	132606
Phone:	973-256-7519
City :	Woodland Park
Zip :	07424
Location :	192 Lackawanna Avenue
County :	Passaic
Municipality :	Woodland Park
Facility :	Waste Management of NJ, Inc.
DEP ID :	448412
Phone:	609-567-4477
City :	Ewing
Zip :	08628
Location :	Route 54 & 5th
County :	Atlantic
Municipality :	Buena Vista
Facility :	Waste Management, Inc.
DEP ID :	194855
Phone:	908-351-2406
City :	Elizabeth
Zip :	07201
Location :	629 S. Front Street
County :	Union
Municipality :	Elizabeth
Facility :	Winzinger, Robert T.
DEP ID :	133555
Phone:	609-267-8600
City :	Hainesport
Zip :	08036
Location :	2989 Delsea Drive
County :	Gloucester
Municipality :	Franklin
"""
    #example of massive string this can parse through and create an Excel file out of
    columnArrOfDict = ColumnSort(massiveString)
    print(columnArrOfDict)
    importData(columnArrOfDict)
 
